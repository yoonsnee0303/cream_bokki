#get data from 도매꾹

# Initializing
import json
import requests
import pandas as pd
import openpyxl as xls
import time

# 도매매 상품리스트
from get_product_numbers import get_product_numbers

# Enter your login credentials and the domeggok_url of the page to retrieve product numbers from
username = 'dfgagu'
password = 'df1051184!@'
domeme_url = 'https://domemedb.domeggook.com/index/item/safeDbList.php?pageLimit=1000' 



# Call the function to retrieve the product numbers
product_numbers = get_product_numbers(username, password, domeme_url)


# Setting domeggok_url
domeggok_url = 'https://domeggook.com/ssl/api/'

print(len(product_numbers))
print(len(product_numbers))
print(len(product_numbers))
print(len(product_numbers))


for num in product_numbers:

# Setting Request Parameters
    params = {
        'ver': '4.4',
        'mode': 'getItemView',
        'aid': 'd761ff40c92e29464d2cc34351dae62a',
        'no': num,
        'om': 'json'
    }

# Getting API Response
    res = requests.get(domeggok_url, params=params)

    pass_point = False

    # Parsing JSON response
    data = json.loads(res.content)

# # Open the JSON file
# with open('data.json',encoding='utf-8') as file:
# # Load the contents of the file
#     data = json.load(file)


    err_cnt = 0
    cnt = 0






    # 상품명, 판매업체, 키워드, 가격, 이미지명 등 데이터 넣기
    dicts = {
    1: data['domeggook']['basis']['title'],                                                                                      # 상품명
    21: data['domeggook']['seller']['nick'],                                                                                     # 판매업체
    11: data['domeggook']['basis']['keywords']['kw'],                                                                            # 키워드
    79: data['domeggook']['price']['dome'],                                                                                      # 가격
    107: data['domeggook']['thumb']['original'],                                                                                 # 이미지명
    }


    # 빈칸은 공백으로 처리
    temp_list = []
    for i in range(149):
        temp_list.append('')


    # 엑셀파일의 컬럼명과 일치하면 해당하는 행에 데이터 넣기.
    for i in range(149):
        for k,v in dicts.items():
            if k == i:
                temp_list[i] = v
                break
    print(temp_list)





    # 나머지 값 채워넣기
    wb = xls.load_workbook(r'C:\Users\Data2\OneDrive\바탕 화면\윤\코드\dom_to_godo\dom_to_godo_rev1.xlsx')
    ws = wb.active
    row = 6                                                                                                                   # 6행부터 데이터 넣을 예정
    column = 1


    # 딕셔너리 내에 리스트가 있을 경우 요소 꺼내기
    for value in temp_list:
        if isinstance(value, list): # value가 list인지 확인
            value = ",".join(map(str,value)) # 리스트들 안에 있는 원소들을 있는 쉼표로 구분하여 출력  
        ws.cell(row=row, column=column).value = value


        if value == '':
            ws.cell(row=row, column=column).value = ws.cell(row=row-1, column=column).value 
        column += 1



    # # Save the changes to the Excel file
    wb.save('C:/Users/Data2/OneDrive/바탕 화면/윤/코드/dom_to_godo/dom_to_godo_rev1.xlsx')



















